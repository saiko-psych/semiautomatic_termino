# -*- coding: utf-8 -*-
"""
tools/update_unicloud_template.py
=================================

Updates an existing versuchsleiter.xlsx in uniCLOUD with fictional test
data: 5 fake supervisors (all using your own Uni email so you receive
all the test mails) and 3 future test termin slots that are far enough
away that no reminder mails go out immediately.

Safety-first:
  * Downloads the current file first, captures its ETag.
  * Saves a local backup at templates/versuchsleiter.backup.<timestamp>.xlsx
    BEFORE any upload attempt.
  * Re-uploads with If-Match: <etag> -> server refuses if the file on
    uniCLOUD changed since our download. No "blind overwrite".
  * Never deletes anything.

Run from the project root:
    python tools/update_unicloud_template.py
"""

from __future__ import annotations

import sys
import shutil
from datetime import datetime, timedelta
from pathlib import Path

# Allow running from the repo root.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("openpyxl is not installed. Run:  pip install -r requirements.txt")
    sys.exit(1)

from utils.unicloud import (
    UniCloudClient,
    UniCloudError,
    UniCloudAuthError,
    UniCloudNotFound,
)

NEXTCLOUD_USERNAME = "your-username_edu"
REMOTE_XLSX = "/Termino/test/versuchsleiter.xlsx"

# All fictional supervisors get THIS email - your own. The script will
# happily mail itself when testing.
TEST_EMAIL = "your-mail@example.org"

# 5 fictional supervisor codes - matches the VL1..VL4 columns in the
# Zeittabelle.
FICTIONAL_VLS = [
    ("TST1", "Test Versuchsleiter 1", TEST_EMAIL),
    ("TST2", "Test Versuchsleiter 2", TEST_EMAIL),
    ("TST3", "Test Versuchsleiter 3", TEST_EMAIL),
    ("TST4", "Test Versuchsleiter 4", TEST_EMAIL),
    ("TST5", "Test Versuchsleiter 5", TEST_EMAIL),
]


def _future_date(days_from_today: int) -> str:
    return (datetime.now() + timedelta(days=days_from_today)).strftime("%d.%m.%Y")


# 3 fictional termin slots, all on the same day, far enough in the future
# (7 days) so today's reminder pass won't fire them.
FUTURE_DAY = _future_date(7)
FAR_FUTURE_DAY = _future_date(14)

FICTIONAL_TIMETABLE = [
    # Datum, Uhrzeit, VL1, VL2, VL3, VL4, Anmerkung VL
    [FUTURE_DAY,      "09:00", "TST1", "TST2", "",     "", "fiktiver Test-Termin"],
    ["",              "10:00", "TST1", "",     "",     "", ""],
    ["",              "15:00", "TST3", "TST2", "TST4", "", ""],
    [FAR_FUTURE_DAY,  "09:00", "TST5", "",     "",     "", "weiterer Test-Termin"],
]


def write_new_workbook(local_path: Path) -> None:
    """Build a fresh xlsx with the fictional data."""
    wb = Workbook()

    # Sheet 1: Zeittabelle
    ws = wb.active
    ws.title = "Zeittabelle"
    headers = ["Datum", "Uhrzeit", "VL1", "VL2", "VL3", "VL4", "Anmerkung VL"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    for row in FICTIONAL_TIMETABLE:
        ws.append(row)
    for col, width in zip("ABCDEFG", [12, 10, 8, 8, 8, 8, 45]):
        ws.column_dimensions[col].width = width

    # Sheet 2: information
    ws_info = wb.create_sheet("information")
    ws_info.append(["VL", "name", "email"])
    for cell in ws_info[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    for vl in FICTIONAL_VLS:
        ws_info.append(list(vl))
    ws_info.column_dimensions["A"].width = 6
    ws_info.column_dimensions["B"].width = 24
    ws_info.column_dimensions["C"].width = 36

    wb.save(local_path)


def main() -> int:
    print("=" * 60)
    print(" uniCLOUD template UPDATE - safety mode")
    print("=" * 60)
    print()
    print(f"Remote file: {REMOTE_XLSX}")
    print(f"Fiktive VLs : 5 (alle mit Mail {TEST_EMAIL})")
    print(f"Test-Tage   : {FUTURE_DAY}, {FAR_FUTURE_DAY} (in 7 / 14 Tagen)")
    print()
    confirm = input("Datei in uniCLOUD ueberschreiben? (j/N): ").strip().lower()
    if confirm not in ("j", "ja", "y", "yes"):
        print("Abgebrochen, nichts gemacht.")
        return 0
    print()

    # 1) connect
    try:
        client = UniCloudClient(username=NEXTCLOUD_USERNAME)
    except UniCloudAuthError as e:
        print(f"AUTH SETUP: {e}")
        return 1

    # 2) download current + ETag (so we can If-Match later)
    print(f"1) Downloading current {REMOTE_XLSX} ...")
    local_dl = Path("templates") / "versuchsleiter.current.xlsx"
    local_dl.parent.mkdir(parents=True, exist_ok=True)
    try:
        current_etag = client.download(REMOTE_XLSX, local_dl)
    except UniCloudNotFound:
        print(f"   FAIL: file not found at {REMOTE_XLSX}. Run "
              f"tools/create_unicloud_template.py first.")
        return 1
    except UniCloudError as e:
        print(f"   FAIL: {e}")
        return 1
    print(f"   ok: {local_dl.stat().st_size} bytes, etag={current_etag!r}")

    # 3) backup
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = Path("templates") / f"versuchsleiter.backup.{ts}.xlsx"
    shutil.copy(local_dl, backup)
    print(f"\n2) Backup saved: {backup}")

    # 4) build new workbook locally
    new_local = Path("templates") / "versuchsleiter.new.xlsx"
    write_new_workbook(new_local)
    print(f"\n3) Built new local workbook: {new_local} "
          f"({new_local.stat().st_size} bytes)")
    print(f"   - 'Zeittabelle': {len(FICTIONAL_TIMETABLE)} Termin-Zeilen")
    print(f"   - 'information': {len(FICTIONAL_VLS)} VL-Stammdaten")

    # 5) upload with If-Match
    print(f"\n4) Uploading with If-Match: {current_etag!r} "
          f"(server refuses if file changed)")
    try:
        new_etag = client.upload(new_local, REMOTE_XLSX, if_match=current_etag)
    except UniCloudError as e:
        print(f"   FAIL: {e}")
        print(f"   Your backup is intact at: {backup}")
        return 1
    print(f"   ok: uploaded (new etag={new_etag!r})")

    # cleanup intermediate file (backup stays)
    new_local.unlink(missing_ok=True)
    local_dl.unlink(missing_ok=True)

    print()
    print("=" * 60)
    print(" DONE")
    print("=" * 60)
    print(f"Remote: {REMOTE_XLSX}  (jetzt mit fiktiven Daten)")
    print(f"Backup: {backup}")
    print()
    print("Next: cloud.uni-graz.at oeffnen, Datei kontrollieren.")
    print("      Dann nochmal `.\\setup_ews.ps1` - es sollten 5 VL-Mails")
    print(f"      (jeweils an {TEST_EMAIL}) verschickt werden, sobald einer der")
    print(f"      Test-Tage ({FUTURE_DAY} / {FAR_FUTURE_DAY}) der nächste Werktag ist.")
    print()
    print("WICHTIG: die 2 Phantom-Termine 01.06.2026 in Termino solltest du")
    print("         manuell loeschen (Browser bei termino.gv.at).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
