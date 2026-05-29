# -*- coding: utf-8 -*-
"""
tools/create_unicloud_template.py
=================================

Creates the supervisor-roster xlsx template for the Termino script and
uploads it to uniCLOUD. Safety-first - this script:

  * does NOT call delete() anywhere (you can grep this file)
  * uses upload_if_new() with If-None-Match: * -> the SERVER refuses if
    the file is already there, no overwriting whatever the local code does
  * checks each parent directory with exists() before mkcol(), and mkcol
    itself is harmless (405 if it already exists)
  * touches exactly two paths:
       1. the directory you choose (created only if missing)
       2. the file <chosen_dir>/versuchsleiter.xlsx (created only if missing)

Run from the project root:
    python tools/create_unicloud_template.py

You will be asked for the target directory. Example: /Termino/Musikstudie

The xlsx has two sheets matching what utils.extensions.google_dp() reads:

  * "Termine"      columns: Datum, Uhrzeit, VL1, VL2, VL3, VL4
  * "information"  columns: VL, name, email

A handful of example rows is included so you can see the layout.
Delete them in Excel/Collabora and add your real data.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

# Allow running from the repo root.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("openpyxl is not installed. Run:  pip install -r requirements.txt")
    sys.exit(1)

from utils.unicloud import (
    UniCloudClient,
    UniCloudError,
    UniCloudAuthError,
)


# Default Nextcloud username. Adjust if your account name is different.
NEXTCLOUD_USERNAME = "your-username_edu"

FILE_NAME = "versuchsleiter.xlsx"


def build_template_xlsx(local_path: Path) -> None:
    """
    Create an xlsx file at ``local_path`` with two sheets matching the
    columns google_dp() expects.
    """
    wb = Workbook()

    # ----- Sheet 1: Zeittabelle (matches the existing Google Sheet layout) -----
    ws = wb.active
    ws.title = "Zeittabelle"
    headers = ["Datum", "Uhrzeit", "VL1", "VL2", "VL3", "VL4", "Anmerkung VL"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # Two demo rows so the layout is visible. Replace with real data after
    # opening the file. Datum is forward-fill: leave empty in following rows
    # that share the same date. VL columns use the abbreviations from the
    # "information" sheet (e.g. LEN, DAV, AMR ...).
    example_rows = [
        ["01.06.2026", "08:00", "XXX", "YYY", "", "", "(Beispielzeile - bitte ersetzen)"],
        ["",            "15:00", "XXX", "",    "", "", ""],
    ]
    for row in example_rows:
        ws.append(row)

    # Reasonable column widths
    for col, width in zip("ABCDEFG", [12, 10, 8, 8, 8, 8, 45]):
        ws.column_dimensions[col].width = width

    # ----- Sheet 2: information (VL-Stammdaten) -----
    ws_info = wb.create_sheet("information")
    info_headers = ["VL", "name", "email"]
    ws_info.append(info_headers)
    for cell in ws_info[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # Single placeholder row. Replace with the real VL list after opening
    # the file in Excel/Collabora. Codes must match what you put into the
    # VL1-VL4 columns above.
    ws_info.append(["XXX", "Beispiel Name", "beispiel@example.org"])

    ws_info.column_dimensions["A"].width = 6
    ws_info.column_dimensions["B"].width = 22
    ws_info.column_dimensions["C"].width = 35

    wb.save(local_path)


def ensure_directory(client: UniCloudClient, directory: str) -> None:
    """
    Create the target directory tree (one level at a time) only where
    needed. NEVER deletes.
    """
    parts = [p for p in directory.strip("/").split("/") if p]
    current = ""
    for part in parts:
        current = f"{current}/{part}"
        if client.exists(current):
            print(f"  ok: {current} already exists")
        else:
            print(f"  creating: {current}")
            client.mkcol(current)


def main() -> int:
    print("=" * 60)
    print(" uniCLOUD template generator - SAFETY MODE")
    print("=" * 60)
    print()
    print("This will:")
    print("  1) ask for a target directory in your uniCLOUD")
    print("  2) create the directory if it does not exist")
    print("  3) build the xlsx LOCALLY and upload it ONLY if")
    print(f"     '<dir>/{FILE_NAME}' does not already exist")
    print()
    print("It will NEVER overwrite or delete anything. If anything would")
    print("be overwritten, the script aborts with an error.")
    print()

    # ---- 1) connect ----
    try:
        client = UniCloudClient(username=NEXTCLOUD_USERNAME)
    except UniCloudAuthError as e:
        print(f"AUTH SETUP: {e}")
        return 1

    # ---- 2) target dir ----
    target_dir = input(
        "Target directory in uniCLOUD\n"
        "(e.g. /Termino/Musikstudie - it will be created if missing): "
    ).strip()
    if not target_dir:
        print("No directory given - aborting.")
        return 1
    if not target_dir.startswith("/"):
        target_dir = "/" + target_dir
    target_dir = target_dir.rstrip("/")
    if not target_dir:
        print("Refusing to operate on root. Pick a subdirectory.")
        return 1

    remote_xlsx = f"{target_dir}/{FILE_NAME}"

    # ---- 3) pre-flight: refuse if file already exists ----
    print(f"\nChecking remote: {remote_xlsx}")
    try:
        if client.exists(remote_xlsx):
            print(f"ABORT: {remote_xlsx} already exists. Not touching it.")
            print(f"       If you want a fresh template, move/rename the existing")
            print(f"       file in uniCLOUD first, then re-run this script.")
            return 1
        print("  ok: target file does not exist yet")
    except UniCloudError as e:
        print(f"FAIL  pre-check error: {e}")
        return 1

    # ---- 4) ensure directory tree ----
    print(f"\nEnsuring directory tree: {target_dir}")
    try:
        ensure_directory(client, target_dir)
    except UniCloudError as e:
        print(f"FAIL  mkcol error: {e}")
        return 1

    # ---- 5) build xlsx locally ----
    local_xlsx = Path("templates") / FILE_NAME
    local_xlsx.parent.mkdir(parents=True, exist_ok=True)
    print(f"\nBuilding local template: {local_xlsx}")
    build_template_xlsx(local_xlsx)
    print(f"  ok: {local_xlsx.stat().st_size} bytes")

    # ---- 6) upload, with server-enforced precondition ----
    print(f"\nUploading to uniCLOUD (If-None-Match: * - won't overwrite):")
    try:
        etag = client.upload_if_new(local_xlsx, remote_xlsx)
    except UniCloudError as e:
        print(f"FAIL  {e}")
        return 1
    print(f"  ok: uploaded (etag={etag})")

    print()
    print("=" * 60)
    print(" DONE")
    print("=" * 60)
    print(f"Local copy : {local_xlsx}")
    print(f"Remote path: {remote_xlsx}")
    print()
    print("Next steps:")
    print(f"  - open {remote_xlsx} in the web (cloud.uni-graz.at) or with")
    print(f"    the Nextcloud desktop client, replace the example rows")
    print(f"    with your real Versuchsleiter data and the actual termin slots.")
    print(f"  - open {remote_xlsx} in the web (cloud.uni-graz.at) or with")
    print(f"    the Nextcloud desktop client, replace the example rows with")
    print(f"    your real Versuchsleiter data and the actual termin slots.")
    print(f"  - then tell me the path '{remote_xlsx}' and I'll wire it into")
    print(f"    the script as a Google-Sheets replacement (sheet_provider).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
